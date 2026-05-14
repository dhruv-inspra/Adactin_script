from __future__ import annotations

import re
import json
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
from pypdf import PdfReader

from .models import Candidate, JobDescription, Question


SUPPORTED_CV_EXTENSIONS = {".docx", ".pdf"}
PHONE_RE = re.compile(r"(?:\+?\d[\d\s().-]{7,}\d)")
EMAIL_RE = re.compile(r"[\w.+-]+@[\w.-]+\.\w+")
WORD_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
SCREENING_LABELS = [
    ("Work Rights", "q004"),
    ("Opportunity Type", "q005"),
    ("Offers in Hand", "q006"),
    ("Notice Period", "q007"),
    ("Total Experience", "q008"),
    ("Guidewire Modules", "q011"),
    ("Configuration & Integration", "q013"),
    ("Gosu Programming", "q014"),
    ("Current Salary", "q015"),
    ("Salary Expectation", "q016"),
]


def parse_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ET.fromstring(xml)
    paragraphs: list[str] = []
    for paragraph in root.findall(".//w:p", WORD_NS):
        text = "".join(t.text or "" for t in paragraph.findall(".//w:t", WORD_NS)).strip()
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def parse_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def parse_document_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".docx":
        return parse_docx_text(path)
    if suffix == ".pdf":
        return parse_pdf_text(path)
    raise ValueError(f"Unsupported document type: {path}")


def normalize_phone(raw: str) -> str:
    digits = re.sub(r"\D", "", raw)
    if not digits:
        return ""
    if digits.startswith("00"):
        digits = digits[2:]
    if digits.startswith("61"):
        return f"+{digits}"
    if digits.startswith("0") and len(digits) >= 9:
        return f"+61{digits[1:]}"
    if raw.strip().startswith("+"):
        return f"+{digits}"
    return f"+{digits}"


def extract_phone(text: str) -> str | None:
    for match in PHONE_RE.findall(text):
        phone = normalize_phone(match)
        if len(re.sub(r"\D", "", phone)) >= 10:
            return phone
    return None


def extract_email(text: str) -> str | None:
    match = EMAIL_RE.search(text)
    return match.group(0) if match else None


def candidate_name_from_file(path: Path) -> str:
    stem = path.stem
    stem = re.sub(r"(?i)[_\s-]*cv$", "", stem)
    stem = stem.replace("_", " ").strip()
    return " ".join(part[:1].upper() + part[1:] for part in stem.split())


def summarize_text(text: str, max_chars: int = 700) -> str:
    compact = re.sub(r"\s+", " ", text).strip()
    if len(compact) <= max_chars:
        return compact
    return compact[: max_chars - 1].rstrip() + "..."


def parse_candidates_from_folder(folder: Path) -> list[Candidate]:
    manifest = load_drive_manifest(folder)
    candidates: list[Candidate] = []
    for path in sorted(folder.iterdir(), key=lambda p: p.name.lower()):
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_CV_EXTENSIONS:
            continue
        if "jd" in path.name.lower():
            continue
        text = parse_document_text(path)
        phone = extract_phone(text)
        if not phone:
            continue
        candidates.append(
            Candidate(
                candidate_id=make_candidate_id(path),
                name=candidate_name_from_file(path),
                phone=phone,
                source_file=path,
                text=text,
                summary=summarize_text(text),
                email=extract_email(text),
                drive_file_id=manifest.get(path.name, {}).get("id"),
                screening_facts=extract_screening_answers_from_cv(text),
            )
        )
    return candidates


def make_candidate_id(path: Path) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "-", path.stem.lower()).strip("-")
    return f"candidate-{cleaned}"


def parse_jd(path: Path) -> JobDescription:
    manifest = load_drive_manifest(path.parent)
    text = parse_document_text(path)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    title = "Guidewire Developer"
    for idx, line in enumerate(lines):
        if line.lower() == "job title" and idx + 1 < len(lines):
            title = lines[idx + 1]
            break
    summary = summarize_jd(lines)
    return JobDescription(
        source_file=path,
        title=title,
        text=text,
        summary=summary,
        drive_file_id=manifest.get(path.name, {}).get("id"),
    )


def load_drive_manifest(folder: Path) -> dict[str, dict[str, str]]:
    path = folder / ".drive_manifest.json"
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def extract_screening_answers_from_cv(text: str) -> dict[str, str]:
    compact = re.sub(r"\s+", " ", text).strip()
    lower = compact.lower()
    facts: dict[str, str] = {}

    name = _first_nonempty_line(text)
    if name:
        facts["q001"] = name
    location = _extract_location(compact)
    if location:
        facts["q002"] = location
    email = extract_email(text)
    if email:
        facts["q003"] = email

    summary_start = lower.find("screening summary")
    summary_compact = compact[summary_start:] if summary_start >= 0 else compact
    summary_lower = summary_compact.lower()
    label_positions: list[tuple[int, str, str]] = []
    for label, question_id in SCREENING_LABELS:
        idx = summary_lower.find(label.lower())
        if idx >= 0:
            label_positions.append((idx, label, question_id))
    label_positions.sort()

    for pos, (idx, label, question_id) in enumerate(label_positions):
        start = idx + len(label)
        end = label_positions[pos + 1][0] if pos + 1 < len(label_positions) else _screening_summary_end(summary_lower, start)
        value = summary_compact[start:end].strip(" :-|")
        if value:
            facts[question_id] = value

    if "q012" not in facts:
        first_section = compact[:500].lower()
        facts["q012"] = "Yes" if "guidewire" in first_section else "No"

    return facts


def _first_nonempty_line(text: str) -> str:
    for line in text.splitlines():
        cleaned = line.strip()
        if cleaned:
            return cleaned
    return ""


def _extract_location(text: str) -> str:
    match = re.search(r"\b([A-Z][A-Za-z]+,\s+[A-Z]{2,3})\s*\|", text)
    return match.group(1) if match else ""


def _screening_summary_end(lower_text: str, start: int) -> int:
    candidates = [idx for marker in ["professional profile", "profile", "experience"] if (idx := lower_text.find(marker, start)) >= 0]
    return min(candidates) if candidates else len(lower_text)


def summarize_jd(lines: list[str]) -> str:
    keep: list[str] = []
    in_required = False
    for line in lines:
        lower = line.lower()
        if lower in {"highly desirable", "why join us", "how to apply"}:
            break
        if lower in {"role overview", "about the role", "essential", "candidate screening criteria"}:
            in_required = True
            continue
        if in_required:
            keep.append(line)
    return summarize_text(" ".join(keep), max_chars=1200)


def parse_questionnaire(path: Path) -> list[Question]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook["Guidewire Developer"] if "Guidewire Developer" in workbook.sheetnames else workbook.active
    header_row = _find_question_header(sheet)
    questions: list[Question] = []
    current: Question | None = None

    for row in range(header_row + 1, sheet.max_row + 1):
        serial = sheet.cell(row, 2).value
        question_text = _clean_cell(sheet.cell(row, 3).value)
        qualifying = _clean_cell(sheet.cell(row, 4).value)

        if _is_question_serial(serial) and question_text:
            current = Question(
                id=f"q{int(serial):03d}",
                text=question_text,
                qualifying_rules=[qualifying] if qualifying else [],
            )
            questions.append(current)
            continue

        if current and qualifying and not question_text:
            current.qualifying_rules.append(qualifying)

    return questions


def _find_question_header(sheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
    for row in range(1, sheet.max_row + 1):
        values = [_clean_cell(sheet.cell(row, col).value).lower() for col in range(1, sheet.max_column + 1)]
        if "questions" in values and "qualifying answers" in values:
            return row
    raise ValueError("Could not find questionnaire header row")


def _is_question_serial(value: object) -> bool:
    if isinstance(value, int):
        return True
    if isinstance(value, float) and value.is_integer():
        return True
    return False


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())
